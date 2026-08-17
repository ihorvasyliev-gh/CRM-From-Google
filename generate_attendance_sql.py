import openpyxl
import re
from datetime import datetime

wb = openpyxl.load_workbook('MANUAL HANDLING 2026.xlsx', data_only=True)

month_map = {
    'jan': 1, 'january': 1,
    'feb': 2, 'february': 2,
    'mar': 3, 'march': 3,
    'apr': 4, 'april': 4,
    'may': 5,
    'jun': 6, 'june': 6,
    'jul': 7, 'july': 7,
    'aug': 8, 'august': 8,
    'sep': 9, 'sept': 9, 'september': 9,
    'oct': 10, 'october': 10,
    'nov': 11, 'november': 11,
    'dec': 12, 'december': 12
}

def parse_date(sheet_name, row2_val):
    text = str(sheet_name) + " " + str(row2_val or "")
    m = re.search(r'(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)\s+(\d{4})', text)
    if m:
        day = int(m.group(1))
        mon_str = m.group(2).lower()
        year = int(m.group(3))
        for k, v in month_map.items():
            if mon_str.startswith(k):
                return f"{year:04d}-{v:02d}-{day:02d}"
    return None

def clean_email(raw_email):
    if not raw_email:
        return ''
    raw = str(raw_email).strip()
    m = re.search(r'[\w\.-]+@[\w\.-]+\.\w+', raw)
    if m:
        return m.group(0).lower()
    return ''

def clean_phone_digits(raw_phone):
    if not raw_phone:
        return ''
    digits = re.sub(r'\D', '', str(raw_phone))
    return digits

attended_records = []

for sheet_name in wb.sheetnames:
    if sheet_name in ['Calendar 2026', 'Sheet1', 'Sheet4', 'Sheet6', 'Sheet330', 'Sheet797']:
        continue
    sheet = wb[sheet_name]
    row2 = sheet.cell(2, 1).value
    sheet_date = parse_date(sheet_name, row2)
    
    for r in range(1, sheet.max_row + 1):
        name = sheet.cell(r, 2).value
        phone = sheet.cell(r, 3).value
        email = sheet.cell(r, 4).value
        referral = sheet.cell(r, 5).value
        group_text = sheet.cell(r, 7).value
        status = sheet.cell(r, 8).value
        
        if name and str(name).strip() not in ['NAME', 'None', 'name'] and not str(name).strip().startswith(('Tuesday', 'Course', 'Manual', 'Monday', 'Thursday', 'Training', 'Name', 'None', 'Sheet')):
            n = str(name).strip()
            st = str(status).strip() if status else ''
            st_low = st.lower()
            
            is_attended = False
            if ('attend' in st_low and 'not' not in st_low and 'didn' not in st_low and 'no' not in st_low) or st_low == 'yes':
                is_attended = True
            
            if is_attended:
                rec_date = sheet_date
                if not rec_date and group_text:
                    m = re.search(r'(\d{1,2})[/\.](\d{1,2})[/\.](\d{2,4})', str(group_text))
                    if m:
                        d, m_val, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
                        if y < 100: y += 2000
                        rec_date = f"{y:04d}-{m_val:02d}-{d:02d}"
                
                n_clean = ' '.join(n.split())
                name_parts = n_clean.split(' ', 1)
                first_name = name_parts[0] if name_parts else ''
                last_name = name_parts[1] if len(name_parts) > 1 else ''
                
                ph_digits = clean_phone_digits(phone)
                em_clean = clean_email(email)
                
                attended_records.append({
                    'sheet': sheet_name,
                    'full_name': n_clean,
                    'first_name': first_name,
                    'last_name': last_name,
                    'phone_raw': str(phone).strip() if phone else '',
                    'phone_digits': ph_digits,
                    'email': em_clean,
                    'date': rec_date or '2026-06-01',
                    'status_raw': st
                })

# Deduplicate
unique_map = {}
for rec in attended_records:
    if rec['email']:
        key = 'em:' + rec['email']
    elif rec['phone_digits'] and len(rec['phone_digits']) >= 7:
        key = 'ph:' + rec['phone_digits'][-7:]
    else:
        key = 'nm:' + rec['full_name'].lower()
    
    if key not in unique_map or (rec['date'] and rec['date'] > unique_map[key]['date']):
        unique_map[key] = rec

unique_records = list(unique_map.values())

sql_lines = []
sql_lines.append("-- ===========================================================================")
sql_lines.append("-- SQL SCRIPT: Update Manual Handling Attendance from MANUAL HANDLING 2026.xlsx")
sql_lines.append("-- Generated: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
sql_lines.append("-- Policy: ONLY update existing students. No new student profiles are created.")
sql_lines.append("-- Pure CTE: No temporary tables created (bypasses Supabase RLS warnings).")
sql_lines.append("-- ===========================================================================\n")

sql_lines.append("WITH raw_attendees (full_name, first_name, last_name, email, phone_digits, completed_date, sheet_source) AS (")
sql_lines.append("    VALUES")

values_strs = []
for r in unique_records:
    fn_esc = r['first_name'].replace("'", "''")
    ln_esc = r['last_name'].replace("'", "''")
    full_esc = r['full_name'].replace("'", "''")
    em_esc = r['email'].replace("'", "''")
    ph_d = r['phone_digits']
    dt = r['date']
    sh_esc = r['sheet'].replace("'", "''")
    values_strs.append(f"    ('{full_esc}', '{fn_esc}', '{ln_esc}', '{em_esc}', '{ph_d}', '{dt}'::date, '{sh_esc}')")

sql_lines.append(",\n".join(values_strs))
sql_lines.append("),")

sql_lines.append("""mh_course AS (
    SELECT id FROM courses WHERE lower(trim(name)) = 'manual handling' LIMIT 1
),
matched_students AS (
    SELECT DISTINCT ON (s.id)
        s.id AS student_id,
        s.first_name AS db_first_name,
        s.last_name AS db_last_name,
        s.email AS db_email,
        s.phone AS db_phone,
        att.full_name AS excel_name,
        att.email AS excel_email,
        att.phone_digits AS excel_phone,
        att.completed_date,
        att.sheet_source,
        CASE 
            WHEN att.email <> '' AND lower(trim(s.email)) = lower(trim(att.email)) THEN 'Email match'
            WHEN length(att.phone_digits) >= 7 AND length(regexp_replace(COALESCE(s.phone, ''), '\\D', '', 'g')) >= 7 
                 AND right(regexp_replace(s.phone, '\\D', '', 'g'), 7) = right(att.phone_digits, 7) THEN 'Phone match'
            WHEN att.first_name <> '' AND att.last_name <> '' 
                 AND lower(trim(s.first_name)) = lower(trim(att.first_name)) 
                 AND lower(trim(s.last_name)) = lower(trim(att.last_name)) THEN 'Exact Name match'
            WHEN att.first_name <> '' AND att.last_name <> '' 
                 AND lower(trim(s.first_name)) = lower(trim(att.last_name)) 
                 AND lower(trim(s.last_name)) = lower(trim(att.first_name)) THEN 'Flipped Name match'
            ELSE 'Other'
        END AS match_reason
    FROM raw_attendees att
    JOIN students s ON (
        (att.email <> '' AND lower(trim(s.email)) = lower(trim(att.email)))
        OR (length(att.phone_digits) >= 7 AND length(regexp_replace(COALESCE(s.phone, ''), '\\D', '', 'g')) >= 7 
            AND right(regexp_replace(s.phone, '\\D', '', 'g'), 7) = right(att.phone_digits, 7))
        OR (att.first_name <> '' AND att.last_name <> '' 
            AND lower(trim(s.first_name)) = lower(trim(att.first_name)) 
            AND lower(trim(s.last_name)) = lower(trim(att.last_name)))
        OR (att.first_name <> '' AND att.last_name <> '' 
            AND lower(trim(s.first_name)) = lower(trim(att.last_name)) 
            AND lower(trim(s.last_name)) = lower(trim(att.first_name)))
    )
),
updated_enrollments AS (
    UPDATE enrollments e
    SET 
        status = 'completed',
        completed_date = m.completed_date,
        completed_at = COALESCE(e.completed_at, m.completed_date::timestamptz, now()),
        updated_at = now()
    FROM matched_students m, mh_course c
    WHERE e.student_id = m.student_id
      AND e.course_id = c.id
    RETURNING e.id AS updated_id
),
inserted_enrollments AS (
    INSERT INTO enrollments (student_id, course_id, status, completed_date, completed_at, updated_at)
    SELECT 
        m.student_id,
        c.id,
        'completed',
        m.completed_date,
        COALESCE(m.completed_date::timestamptz, now()),
        now()
    FROM matched_students m
    CROSS JOIN mh_course c
    WHERE NOT EXISTS (
        SELECT 1 FROM enrollments e2 
        WHERE e2.student_id = m.student_id 
          AND e2.course_id = c.id
    )
    RETURNING id AS inserted_id
),
stats AS (
    SELECT 
        (SELECT count(*) FROM updated_enrollments) AS num_updated,
        (SELECT count(*) FROM inserted_enrollments) AS num_new_enrolled
)
SELECT 
    m.student_id,
    m.db_first_name || ' ' || m.db_last_name AS db_full_name,
    m.db_email,
    m.db_phone,
    m.excel_name,
    m.completed_date,
    m.sheet_source,
    m.match_reason,
    (SELECT num_updated FROM stats) AS total_enrollments_updated,
    (SELECT num_new_enrolled FROM stats) AS total_new_enrollments_created
FROM matched_students m
ORDER BY m.completed_date DESC, m.db_last_name ASC;
""")

output_sql = "\n".join(sql_lines)
with open('update_manual_handling_attendance.sql', 'w', encoding='utf-8') as f:
    f.write(output_sql)

print("Updated update_manual_handling_attendance.sql successfully!")
